import pandas as pd
from Data1.GDM.Preprocess.utils import drop_instances_from_mom_list, find_joint_nodes_set
from Mother import *
from MotherCollection import *
import networkx as nx
from Bacteria import *
import os
from openpyxl import load_workbook
import pickle
import matplotlib.pyplot as plt


# function from Ariel Rozen
# def create_tax_tree(series, zeroflag=False):
#     tempGraph = nx.Graph()
#     """workbook = load_workbook(filename="random_Otus.xlsx")
#     sheet = workbook.active"""
#     valdict = {}
#     bac = []
#     for i, (tax, val) in enumerate(series.items()):
#         # adding the bacteria in every column
#         bac.append(Bacteria(tax, val))
#         # connecting to the root of the tempGraph
#         tempGraph.add_edge("anaerobe", (bac[i].lst[0],))
#         # connecting all levels of the taxonomy
#         for j in range(0, len(bac[i].lst) - 1):
#             updateval(tempGraph, bac[i], valdict, j, True)
#         # adding the value of the last node in the chain
#         updateval(tempGraph, bac[i], valdict, len(bac[i].lst) - 1, False)
#     valdict["anaerobe"] = valdict[("Bacteria",)] + valdict[("Archaea",)]
#     return create_final_graph(tempGraph, valdict, zeroflag)
#
#
# def updateval(graph, bac, vald, num, adde):
#     if adde:
#         graph.add_edge(tuple(bac.lst[:num + 1]), tuple(bac.lst[:num + 2]))
#     # adding the value of the nodes
#     if tuple(bac.lst[:num + 1]) in vald:
#         vald[tuple(bac.lst[:num + 1])] += bac.val
#     else:
#         vald[tuple(bac.lst[:num + 1])] = bac.val

def create_tax_tree(series, zeroflag=False):
    tempGraph = nx.Graph()
    valdict = {}
    bac = []
    for i, (tax, val) in enumerate(series.items()):
        # adding the bacteria in every column
        bac.append(Bacteria(tax, val))
        # connecting to the root of the tempGraph
        tempGraph.add_edge("anaerobe", bac[i].lst[0])
        # connecting all levels of the taxonomy
        for j in range(0, len(bac[i].lst) - 1):
            updateval(tempGraph, bac[i], valdict, j, True)
        # adding the value of the last node in the chain
        updateval(tempGraph, bac[i], valdict, len(bac[i].lst) - 1, False)
    valdict["anaerobe"] = valdict["Bacteria"] + valdict["Archaea"]
    return create_final_graph(tempGraph, valdict, zeroflag)


def updateval(graph, bac, vald, num, adde):
    if adde:
        graph.add_edge(bac.lst[num], bac.lst[num + 1])
    # adding the value of the nodes
    if bac.lst[num] in vald:
        vald[bac.lst[num]] += bac.val
    else:
        vald[bac.lst[num]] = bac.val

def create_final_graph(tempGraph, valdict, zeroflag):
    graph = nx.Graph()
    for e in tempGraph.edges():
        if not zeroflag or (valdict[e[0]] != -1 and valdict[e[1]] != -1):
            graph.add_edge((e[0], valdict[e[0]]),
                           (e[1], valdict[e[1]]))
    return graph


def draw_tree(graph, threshold=0.0):
    if type(threshold) == tuple:
        lower_threshold, higher_threshold = threshold
    else:
        lower_threshold, higher_threshold = -threshold, threshold
    labelg = {}
    labelr = {}
    colormap = []
    sizemap = []
    for node in graph:
        if node[0] == "base":
            colormap.append("white")
            sizemap.append(0)
        else:
            if node[1] < lower_threshold:
                colormap.append("red")
                labelr[node] = node[0][-1]
            elif node[1] > higher_threshold:
                colormap.append("green")
                labelg[node] = node[0][-1]
            else:
                colormap.append("yellow")
            sizemap.append(node[1] / 1000 + 5)
    # drawing the graph
    # pos = graphviz_layout(graph, prog="twopi", root="base")
    pos = nx.spring_layout(graph)
    lpos = {}
    for key, loc in pos.items():
        lpos[key] = (loc[0], loc[1] + 0.02)
    nx.draw(graph, pos, node_size=sizemap, node_color=colormap, width=0.3)
    nx.nx.draw_networkx_labels(graph, lpos, labelr, font_size=7, font_color="red")
    nx.nx.draw_networkx_labels(graph, lpos, labelg, font_size=7, font_color="green")
    plt.draw()
    plt.show()
    plt.savefig("taxtree.png")


def return_microbiome_dict(all_moms_dataframe):
    microbiome_dict = {}
    j = 0
    for i, mom in enumerate(all_moms_dataframe.iterrows()):
        cur_graph = create_tax_tree(all_moms_dataframe.iloc[i], zeroflag=True)
        nodes = cur_graph.nodes(data=False)
        for name, value in nodes:
            if name not in microbiome_dict:
                # print(name)
                microbiome_dict[name] = j
                j = j + 1
    return microbiome_dict


def create_graph_files_for_qgcn_new(taxonomy_file_name, mapping_table_file_name, graph_csv_file, external_data_file):
    type_file = '.csv'
    graph_csv_file = open(graph_csv_file + type_file, "wt")
    external_data_file = open(external_data_file + type_file, "wt")
    # header of files
    graph_header = "g_id,src,dst,label\n"
    graph_csv_file.write(graph_header)
    external_data_header = "g_id,node"

    all_moms_dataframe = load_taxonomy_file(taxonomy_file_name, ',')
    mapping_table_dataframe = load_tags_file(mapping_table_file_name)

    # we need to add column for each microbiome, in order to keep our one-hot vector idea
    microbiome_dict = return_microbiome_dict(all_moms_dataframe)
    microbiom_len = len(microbiome_dict)

    microbiom_len = len(microbiome_dict)
    for r in range(microbiom_len):
        external_data_header += f',{r}'
    external_data_header += '\n'
    external_data_file.write(external_data_header)
    # i = 0
    # for microbiome in all_moms_dataframe.columns:
    #     microbiome_dict[microbiome] = i
    #     i += 1
    # microbiome_dict['anaerobe'] = i
    # i = i+1
    # microbiome_dict['Archaea'] = i
    #
    j = 0
    # creation of the two files for qgcn - graph file, and external data graph file
    for i, mom in enumerate(all_moms_dataframe.iterrows()):
        cur_graph = create_tax_tree(all_moms_dataframe.iloc[i], zeroflag=True)
        # draw_tree(cur_graph)
        edges = cur_graph.edges(data=False)
        nodes = cur_graph.nodes(data=False)
        label = mapping_table_dataframe.iloc[i]['Tag']
        for u, v in edges:
            name1, value1 = u
            name2, value2 = v
            # if name1 not in microbiome_dict:
            #     microbiome_dict[name1] = j
            #     j = j + 1
            # if name2 not in microbiome_dict:
            #     microbiome_dict[name2] = j
            #     j = j + 1
            line = f"{i},{microbiome_dict[name1]},{microbiome_dict[name2]},{label}\n"
            graph_csv_file.write(line)
        for name, value in nodes:
            microbiome_id = microbiome_dict[name]
            line = f"{i},{microbiome_id}"
            one_hot = [0 if j != microbiome_id else value for j in range(microbiom_len)]
            line += ''.join(',' + str(e) for e in one_hot)
            line += '\n'
            external_data_file.write(line)
    graph_csv_file.close()
    external_data_file.close()


def iterate_dataframe(all_moms_dataframe, mapping_table_dataframe, save_to_pickle=False):
    mom_list = MotherCollection()
    taxonomy_series = all_moms_dataframe.iloc[-1]
    for index, mom in mapping_table_dataframe.iterrows():
        print("index", index)
        split_sample_id = mom['ID'].split('-')
        id = split_sample_id[0]
        test_way = split_sample_id[1]
        trimester = split_sample_id[2][-1]
        repetition = split_sample_id[-1]
        sick_or_not = mom['Tag']
        data = []
        ind = []
        # for i in range(all_moms_dataframe.shape[1]):
        # microbiome_id = all_moms_dataframe.columns[i]
        # ind.append(taxonomy_series[microbiome_id])
        # data.append(all_moms_dataframe.iloc[index - 1][microbiome_id])
        # my_series = pd.Series(data, index=ind)
        my_series = all_moms_dataframe.iloc[index]
        microbiome_graph = create_tax_tree(my_series, zeroflag=True)
        mom_object = Mother(id, trimester, repetition, sick_or_not, microbiome_graph, test_way=test_way)
        mom_list.add_mom(mom_object)
    if save_to_pickle:
        pickle.dump(mom_list, open("mom_collection.p", "wb"))
    return mom_list


def create_graph_files_for_qgcn(mom_list, graph_csv_file, external_data_file, microbiome_dict):
    type_file = '.csv'
    graph_csv_file = open(graph_csv_file + type_file, "wt")
    external_data_file = open(external_data_file + type_file, "wt")
    # header of files
    graph_header = "g_id,src,dst,label\n"
    graph_csv_file.write(graph_header)
    external_data_header = "g_id,node"
    # we need to add column for each microbiome, in order to keep our one-hot vector idea
    microbiom_len = len(microbiome_dict)
    for i in range(microbiom_len):
        external_data_header += f',{i}'
    external_data_header += '\n'
    external_data_file.write(external_data_header)
    # creation of the two files for qgcn - graph file, and external data graph file
    for i, mom in enumerate(mom_list.mother_list):
        cur_graph = mom.microbiome_graph
        draw_tree(cur_graph)
        edges = cur_graph.edges(data=False)
        nodes = cur_graph.nodes(data=False)
        label = mom.sick_or_not
        for u, v in edges:
            name1, value1 = u
            name2, value2 = v
            line = f"{i},{microbiome_dict[name1]},{microbiome_dict[name2]},{label}\n"
            graph_csv_file.write(line)
        for name, value in nodes:
            microbiome_id = microbiome_dict[name]
            line = f"{i},{microbiome_id}"
            one_hot = [0 if j != microbiome_id else value for j in range(microbiom_len)]
            line += ''.join(',' + str(e) for e in one_hot)
            line += '\n'
            external_data_file.write(line)
    graph_csv_file.close()
    external_data_file.close()


def load_taxonomy_file(file_name, delimeter='\t'):
    if delimeter == '\t':
        all_moms_dataframe = pd.read_csv(file_name, delimiter=delimeter, header=1)
        all_moms_dataframe = all_moms_dataframe.T
        new_header = all_moms_dataframe.iloc[0]  # grab the first row for the header
        all_moms_dataframe = all_moms_dataframe[1:]  # take the data less the header row
        all_moms_dataframe.columns = new_header  # set the header row as the df header
        all_moms_dataframe.index.name = 'ID'
    if delimeter == ',':
        all_moms_dataframe = pd.read_csv(file_name, delimiter=delimeter, header=0)
        all_moms_dataframe.set_index('ID', inplace=True)
    all_moms_dataframe.sort_index(inplace=True)
    return all_moms_dataframe


def load_tags_file(file_name):
    wb = load_workbook(file_name)
    ws = wb['Sheet1']
    data = []
    for row in ws:
        if not ws.row_dimensions[row[0].row].hidden:
            row_values = [cell.value for cell in row]
            data.append(row_values)
    mapping_table_dataframe = pd.DataFrame(data[1:], columns=data[0])
    # mapping_table_dataframe = pd.read_excel(mapping_table, header=0)
    # mapping_table_dataframe = mapping_table_dataframe[1:]
    mapping_table_dataframe.dropna(axis=1, inplace=True)
    mapping_table_dataframe.rename(columns={"#SampleID": "ID", "Control_GDM": "Tag"}, inplace=True)
    mapping_table_dataframe.replace({"Tag": {"GDM": 1, "Control": 0}}, inplace=True)

    mapping_table_dataframe = mapping_table_dataframe[["ID", "Tag", 'trimester']]
    mapping_table_dataframe.set_index("ID", inplace=True)
    mapping_table_dataframe.sort_index(inplace=True)
    return mapping_table_dataframe


def create_moms_list_for_files_for_qgcn(taxonomy_file_name, mapping_table_file_name, ready_pickle=True):
    if not ready_pickle:
        # The taxonomy file is after MIP-MLP site Preprocess
        all_moms_dataframe = load_taxonomy_file(taxonomy_file_name, delimeter=',')
        all_moms_dataframe.to_csv("taxonomy_gdm_file.csv")
        mapping_table_dataframe = load_tags_file(mapping_table_file_name)
        mapping_table_dataframe.to_csv("tag_gdm_file.csv", index=False)
        mom_list = iterate_dataframe(all_moms_dataframe, mapping_table_dataframe, save_to_pickle=False)
        # mom_list = drop_instances_from_mom_list(mom_list)
        # a = 5

    mom_list = pickle.load(open("mom_collection.p", "rb"))

    # find the number of all different nodes in all graphs
    microbiome_dict = find_joint_nodes_set(mom_list)
    mom_list = drop_instances_from_mom_list(mom_list)
    create_graph_files_for_qgcn(mom_list, "microbiome_data_for_qgcn", "external_microbiome_data_for_qgcn",
                                microbiome_dict)


if __name__ == '__main__':
    path_data_dir = os.path.join("..", "israel")
    # OTU_merged_Mucositis.csv file is after MIP-MLP site
    taxonomy_file_name = os.path.join(path_data_dir, "OTU_merged_Mucositis_after_mipmlp_new.csv")
    # taxonomy_file_name = os.path.join(path_data_dir, "OTU_merged_Mucositis.csv")
    mapping_table = os.path.join(path_data_dir, "israeli_stool_mapping_table.xlsx")
    # create_moms_list_for_files_for_qgcn(taxonomy_file_name, mapping_table, ready_pickle=False)
    # microbiome_dict = find_joint_nodes_set(mom_list)
    create_graph_files_for_qgcn_new(taxonomy_file_name, mapping_table, "microbiome_data_for_qgcn",
                                    "external_microbiome_data_for_qgcn")
